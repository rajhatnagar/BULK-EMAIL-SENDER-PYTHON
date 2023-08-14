import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl

def send_email(subject, message, from_email, to_emails, smtp_server, smtp_port, smtp_username, smtp_password):
    try:
        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = ", ".join(to_emails)
        msg['Subject'] = subject

        msg.attach(MIMEText(message, 'plain'))

        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_username, smtp_password)
        server.sendmail(from_email, to_emails, msg.as_string())
        server.quit()

        return True

    except Exception as e:
        print("Error sending email:", e)
        return False

def get_emails_from_excel(filename):
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        email_column = sheet['A']  # Assuming emails are in the first column

        email_list = [cell.value for cell in email_column if cell.value]
        return email_list

    except Exception as e:
        print("Error:", e)
        return []

def update_status_in_excel(filename, email_statuses):
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        status_column = sheet['B']  # Assuming status will be in the second column

        for i, status in enumerate(email_statuses, start=1):
            status_column.cell(row=i, column=2, value=status)

        workbook.save(filename)
        print("Status updated successfully.")

    except Exception as e:
        print("Error updating status:", e)

if __name__ == "__main__":
    subject = "mai aa gaya"
    message = "This is a test email sent from my tool."
    from_email = "data@buydatabase.online"  # Your Titan Email

    smtp_server = "smtp.titan.email"  # Titan Email SMTP server
    smtp_port = 587
    smtp_username = "data@buydatabase.online"  # Your Titan Email
    smtp_password = "#f544Oy1LL^5"  # Your Titan Email password or app-specific password

    excel_filename = "emails.xlsx"
    to_emails = get_emails_from_excel(excel_filename)
    
    email_statuses = []
    for email in to_emails:
        if send_email(subject, message, from_email, [email], smtp_server, smtp_port, smtp_username, smtp_password):
            email_statuses.append("Sent")
        else:
            email_statuses.append("Failed")

    update_status_in_excel(excel_filename, email_statuses)
